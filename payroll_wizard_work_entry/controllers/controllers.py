# -*- coding: utf-8 -*-

import base64
import io
import openpyxl
from odoo import http


class ImportLots(http.Controller):
    """Class to handle excel download"""
    @http.route('/download/inputs_excel', type='http', auth="user")
    def download_inputs_excel_file(self):
        """Download sample Excel sheet"""
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        # Add headers
        headers = [
            ('Empleado', 'Nombre o código del empleado (considerar mayusculas, minusculas y espacios)'),
            ('Variable', 'Nombre o código de la variable de entrada (considerar mayusculas, minusculas y espacios)'),
            ('Monto', 'Monto correspondiente a la variable (numeros enteros o decimales, no negativos)'),
        ]
        for col, (header, comment_text) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            comment = openpyxl.comments.Comment(comment_text, 'System')
            cell.comment = comment
        # Add sample data
        data = [
            ('Mitchel Admin', 'Otras bonificaciones', 2.00),
            ('Marc Demo', 'Otros descuentos', 3.00),
            ('Joel Willis', 'Horas extras', 6.00),
        ]
        for row in data:
            ws.append(row)
        # Save the workbook to a BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        # Convert the buffer content to base64
        file_content_base64 = base64.b64encode(buffer.getvalue())
        return http.send_file(io.BytesIO(base64.b64decode(file_content_base64)),
                              filename='import_inputs_template.xlsx',
                              as_attachment=True,
                              mimetype='application/vnd.'
                                       'openxmlformats-officedocument.'
                                       'spreadsheetml.sheet')
    
    @http.route('/download/workentry_excel', type='http', auth="user")
    def download_workentry_excel_file(self):
        """Download sample Excel sheet"""
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        # Add headers
        headers = [
            ('Empleado', 'Nombre o código del empleado (considerar mayusculas, minusculas y espacios)'),
            ('Tipo de entrada', 'Nombre o código de la variable de entrada (considerar mayusculas, minusculas y espacios)'),
            ('Dias', 'Dias correspondientes a la entrada (numeros enteros o decimales, no negativos)'),
            ('Horas', 'Horas correspondientes a la entrada (numeros enteros o decimales, no negativos)'),
        ]
        for col, (header, comment_text) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            comment = openpyxl.comments.Comment(comment_text, 'System')
            cell.comment = comment
        # Add sample data
        data = [
            ('Mitchel Admin', 'Asistencia', 2.00, 0.00),
            ('Marc Demo', 'Vacaciones', 3.00, 0.00),
            ('Joel Willis', 'Horas extras', 0.00, 6.00),
        ]
        for row in data:
            ws.append(row)
        # Save the workbook to a BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        # Convert the buffer content to base64
        file_content_base64 = base64.b64encode(buffer.getvalue())
        return http.send_file(io.BytesIO(base64.b64decode(file_content_base64)),
                              filename='import_workentry_template.xlsx',
                              as_attachment=True,
                              mimetype='application/vnd.'
                                       'openxmlformats-officedocument.'
                                       'spreadsheetml.sheet')
